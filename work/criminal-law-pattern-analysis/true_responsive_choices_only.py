from __future__ import annotations

import csv
import itertools
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(r"C:\Users\JesusLovesMe\Documents\CrimL\Criminal LAw.xlsx")
OUT = Path(r"C:\ABM\work\criminal-law-pattern-analysis")
LETTERS = "ABCD"

WORD_RE = re.compile(r"[A-Za-z0-9']+")
BECAUSE_RE = re.compile(r"\b(because|since|as|therefore)\b", re.I)
LIMIT_RE = re.compile(r"\b(if|unless|only if|provided|where|when|while|although|except|even if|so long as)\b", re.I)
ABSOLUTE_RE = re.compile(r"\b(always|never|all|none|any|every|automatically|necessarily|under any circumstances|no circumstances)\b", re.I)
COMMAND_RE = re.compile(r"\b(must|cannot|can only|required|requirement|barred|prohibited)\b", re.I)
HEDGE_RE = re.compile(r"\b(generally|usually|may|might|likely|probably|can|could|unless|if)\b", re.I)
BARE_RESULT_RE = re.compile(r"^\s*(yes|no|guilty|not guilty|admissible|inadmissible|reversed|affirmed|granted|denied|grant|deny)\b[\s.,;:!-]*$", re.I)
CONCLUSION_RE = re.compile(r"^\s*(yes|no|guilty|not guilty|admissible|inadmissible|reversed|affirmed|grant|deny|granted|denied)\b", re.I)
NEG_RE = re.compile(r"\b(no|not|neither|without|fails?|invalid|inadmissible|suppress(?:ed|ible)?|barred)\b", re.I)
EXTREME_RE = re.compile(r"\b(solely|merely|simply|completely|entirely|permanently|absolute|strictly)\b", re.I)


def words(text: str) -> list[str]:
    return WORD_RE.findall(text or "")


def wc(text: str) -> int:
    return len(words(text))


def as_num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_records():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    records = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        qid = str(row.get("barmatrix_question_id") or "").strip()
        if not qid:
            continue
        choices = {letter: str(row.get(f"answer_{letter.lower()}") or "").strip() for letter in LETTERS}
        correct = str(row.get("correct_answer") or "").strip().upper()[:1]
        if correct not in LETTERS or not any(choices.values()):
            continue
        records.append(
            {
                "qid": qid,
                "choices": choices,
                "correct": correct,
                "pc": as_num(row.get("percent_correct")),
            }
        )
    return records


def feature_counts(choice: str, all_choices: dict[str, str]):
    lengths = {letter: wc(text) for letter, text in all_choices.items()}
    max_len = max(lengths.values()) or 1
    text = choice or ""
    return {
        "length": wc(text),
        "rel_length": wc(text) / max_len,
        "is_longest": 1 if wc(text) == max_len else 0,
        "because": len(BECAUSE_RE.findall(text)),
        "limited": len(LIMIT_RE.findall(text)),
        "absolute": len(ABSOLUTE_RE.findall(text)),
        "command": len(COMMAND_RE.findall(text)),
        "hedge": len(HEDGE_RE.findall(text)),
        "neg": len(NEG_RE.findall(text)),
        "extreme": len(EXTREME_RE.findall(text)),
        "bare_result": 1 if BARE_RESULT_RE.search(text) else 0,
        "has_conclusion": 1 if CONCLUSION_RE.search(text) else 0,
        "semicolon": text.count(";"),
        "comma": text.count(","),
    }


def tear_score(choice: str, all_choices: dict[str, str], weights: dict[str, float]):
    f = feature_counts(choice, all_choices)
    return sum(weights.get(name, 0.0) * value for name, value in f.items())


def pick_by_score(weights):
    def fn(rec):
        scores = {letter: tear_score(rec["choices"][letter], rec["choices"], weights) for letter in LETTERS}
        max_score = max(scores.values())
        cands = [letter for letter in LETTERS if scores[letter] == max_score]
        if len(cands) == 1:
            return cands[0]
        # Residual tie-break: C is the strongest baseline in this bank.
        return "C" if "C" in cands else cands[0]
    return fn


def pick_longest_abc_else_c(rec):
    lengths = {letter: wc(rec["choices"][letter]) for letter in LETTERS}
    max_len = max(lengths.values())
    cands = [letter for letter in LETTERS if lengths[letter] == max_len]
    if len(cands) == 1 and cands[0] in "ABC":
        return cands[0]
    return "C"


def pick_truth_responsive_v1(rec):
    weights = {
        "rel_length": 4.0,
        "because": 1.0,
        "limited": 0.8,
        "hedge": 0.3,
        "absolute": -1.4,
        "command": -0.5,
        "extreme": -0.8,
        "bare_result": -2.0,
    }
    return pick_by_score(weights)(rec)


def evaluate(records, fn):
    if not records:
        return 0.0
    return sum(1 for rec in records if fn(rec) == rec["correct"]) / len(records)


def folds(n, k=5):
    return {i: i % k for i in range(n)}


def cv_grid_search(records):
    # Keep this small and interpretable. Features are mechanical TRUE/RESPONSIVE
    # proxies, not doctrine terms.
    grid = {
        "rel_length": [3.0, 4.0, 5.0],
        "because": [0.0, 1.0],
        "limited": [0.0, 0.8],
        "absolute": [-1.5, 0.0],
        "bare_result": [-2.0, 0.0],
    }
    fold_of = folds(len(records), 5)
    total_hits = 0
    fold_weights = []
    for fold in range(5):
        train = [rec for i, rec in enumerate(records) if fold_of[i] != fold]
        test = [rec for i, rec in enumerate(records) if fold_of[i] == fold]
        best_acc = -1
        best_weights = None
        keys = list(grid)
        for vals in itertools.product(*(grid[k] for k in keys)):
            weights = dict(zip(keys, vals))
            acc = evaluate(train, pick_by_score(weights))
            if acc > best_acc:
                best_acc = acc
                best_weights = weights
        fold_weights.append({"fold": fold, "train_acc": best_acc, "weights": best_weights})
        fn = pick_by_score(best_weights)
        total_hits += sum(1 for rec in test if fn(rec) == rec["correct"])
    return total_hits / len(records), fold_weights


def bucket_stats(records):
    out = []
    buckets = defaultdict(list)
    for rec in records:
        choices = rec["choices"]
        lengths = {letter: wc(choices[letter]) for letter in LETTERS}
        max_len = max(lengths.values())
        longest = "".join(letter for letter in LETTERS if lengths[letter] == max_len)
        has_because = any(BECAUSE_RE.search(choices[letter]) for letter in LETTERS)
        has_absolute = any(ABSOLUTE_RE.search(choices[letter]) for letter in LETTERS)
        key = (longest, has_because, has_absolute)
        buckets[key].append(rec)
    for key, group in buckets.items():
        correct_counts = Counter(rec["correct"] for rec in group)
        out.append(
            {
                "longest_set": key[0],
                "any_because": key[1],
                "any_absolute": key[2],
                "n": len(group),
                "best_letter": correct_counts.most_common(1)[0][0],
                "best_letter_acc": correct_counts.most_common(1)[0][1] / len(group),
                "longest_abc_else_c_acc": evaluate(group, pick_longest_abc_else_c),
                "truth_responsive_v1_acc": evaluate(group, pick_truth_responsive_v1),
            }
        )
    return sorted(out, key=lambda row: row["n"], reverse=True)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    records = load_records()
    graded = [rec for rec in records if rec["pc"] is not None]
    critical = [rec for rec in records if rec["pc"] is not None and rec["pc"] < 50]

    cv_acc, fold_weights = cv_grid_search(records)
    rules = {
        "longest_A/B/C_else_C": pick_longest_abc_else_c,
        "truth_responsive_v1": pick_truth_responsive_v1,
        "cv_grid_truth_responsive": None,
    }
    scoreboard = []
    for name, fn in rules.items():
        if fn is None:
            continue
        scoreboard.append(
            {
                "rule": name,
                "all_acc": evaluate(records, fn),
                "graded_acc": evaluate(graded, fn),
                "critical_acc": evaluate(critical, fn),
            }
        )
    scoreboard.append(
        {
            "rule": "cv_grid_truth_responsive",
            "all_acc": cv_acc,
            "graded_acc": None,
            "critical_acc": None,
        }
    )
    scoreboard.sort(key=lambda row: row["all_acc"], reverse=True)

    results = {
        "n": len(records),
        "graded_n": len(graded),
        "critical_n": len(critical),
        "scoreboard": scoreboard,
        "cv_fold_weights": fold_weights,
        "bucket_stats": bucket_stats(records),
        "note": "Features are answer-choice-only TRUE/RESPONSIVE proxies: length, because/reason coupling, limited language, absolute/overclaim penalties, and bare-result penalties.",
    }
    (OUT / "true_responsive_choices_only_results.json").write_text(json.dumps(results, indent=2), encoding="utf-8")

    with (OUT / "true_responsive_choices_only_scoreboard.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(scoreboard[0].keys()))
        writer.writeheader()
        writer.writerows(scoreboard)

    buckets = results["bucket_stats"]
    with (OUT / "true_responsive_choices_only_buckets.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(buckets[0].keys()))
        writer.writeheader()
        writer.writerows(buckets)

    print(json.dumps({
        "n": results["n"],
        "scoreboard": scoreboard,
        "cv_fold_weights": fold_weights,
        "top_buckets": buckets[:12],
    }, indent=2))


if __name__ == "__main__":
    main()
