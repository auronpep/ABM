from __future__ import annotations

import csv
import json
import math
import random
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook

SRC = Path(r"C:\Users\JesusLovesMe\Documents\CrimL\Criminal LAw.xlsx")
OUT = Path(r"C:\ABM\work\criminal-law-pattern-analysis")

LETTERS = "ABCD"
WORD_RE = re.compile(r"[A-Za-z0-9']+")
QUAL_RE = re.compile(r"\b(if|because|since|unless|only|provided|where|when|while|although|except|even if)\b", re.I)
ABS_RE = re.compile(r"\b(always|never|all|none|cannot|must|any|every|automatically|necessarily)\b", re.I)
NEG_RE = re.compile(r"\b(no|not|neither|nor|without|fails?|invalid|inadmissible|suppress(?:ed|ible)?)\b", re.I)
REMEDY_RE = re.compile(r"\b(admissible|inadmissible|suppress(?:ed|ible)?|reversed|affirmed|grant(?:ed)?|deny|denied|harmless|violation|motion)\b", re.I)
RIGHT_RE = re.compile(r"\b(fifth|sixth|fourth|amendment|miranda|counsel|confrontation|due process|brady|double jeopardy)\b", re.I)
CRIME_RE = re.compile(r"\b(murder|manslaughter|larceny|embezzlement|pretenses|robbery|burglary|arson|forgery|attempt|conspiracy|solicitation|assault|battery|rape|kidnap|mayhem)\b", re.I)


def words(value: str) -> list[str]:
    return WORD_RE.findall(value or "")


def word_count(value: str) -> int:
    return len(words(value))


def as_num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_records():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}
    records = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() for v in raw):
            continue
        row = dict(zip(headers, raw))
        qid = str(row.get("barmatrix_question_id") or "").strip()
        choices = {letter: str(row.get(f"answer_{letter.lower()}") or "").strip() for letter in LETTERS}
        correct = str(row.get("correct_answer") or "").strip().upper()[:1]
        if correct not in LETTERS or not any(choices.values()):
            continue
        records.append(
            {
                "qid": qid,
                "choices": choices,
                "correct": correct,
                "popular_wrong": str(row.get("most_popular_wrong_answer") or "").strip().upper()[:1],
                "percent_correct": as_num(row.get("percent_correct")),
                "choice_pcts": {letter: as_num(row.get(f"percent_{letter.lower()}")) for letter in LETTERS},
            }
        )
    return records


def choice_features(choice: str, letter: str, all_choices: dict[str, str]) -> dict[str, float]:
    counts = {l: word_count(v) for l, v in all_choices.items()}
    chars = {l: len(v) for l, v in all_choices.items()}
    wc = counts[letter]
    cc = chars[letter]
    max_wc = max(counts.values()) or 1
    max_cc = max(chars.values()) or 1
    sorted_wc = sorted(set(counts.values()), reverse=True)
    sorted_cc = sorted(set(chars.values()), reverse=True)
    length_rank = sorted_wc.index(wc) + 1
    char_rank = sorted_cc.index(cc) + 1
    text = choice or ""
    return {
        "word_count": wc,
        "char_count": cc,
        "rel_word_len": wc / max_wc,
        "rel_char_len": cc / max_cc,
        "is_longest_words": 1.0 if wc == max_wc else 0.0,
        "is_longest_chars": 1.0 if cc == max_cc else 0.0,
        "word_length_rank": float(length_rank),
        "char_length_rank": float(char_rank),
        "qualifiers": float(len(QUAL_RE.findall(text))),
        "absolutes": float(len(ABS_RE.findall(text))),
        "negations": float(len(NEG_RE.findall(text))),
        "remedy_words": float(len(REMEDY_RE.findall(text))),
        "right_words": float(len(RIGHT_RE.findall(text))),
        "crime_words": float(len(CRIME_RE.findall(text))),
        "has_because": 1.0 if re.search(r"\bbecause\b", text, re.I) else 0.0,
        "starts_yes": 1.0 if re.match(r"\s*yes\b", text, re.I) else 0.0,
        "starts_no": 1.0 if re.match(r"\s*no\b", text, re.I) else 0.0,
        "has_guilty": 1.0 if re.search(r"\bguilty\b", text, re.I) else 0.0,
        "has_not_guilty": 1.0 if re.search(r"\bnot guilty\b", text, re.I) else 0.0,
        "has_only": 1.0 if re.search(r"\bonly\b", text, re.I) else 0.0,
        "position": float(LETTERS.index(letter)),
    }


def feature_vector(choice: str, letter: str, all_choices: dict[str, str], names: list[str]) -> list[float]:
    feats = choice_features(choice, letter, all_choices)
    return [feats[name] for name in names]


def menu_type(choices: dict[str, str]) -> str:
    text = " ".join(choices.values()).lower()
    def hits(terms):
        return sum(1 for term in terms if term in text)
    if ("both" in text and "neither" in text) or "proposal a" in text or "proposal b" in text or "all of them" in text:
        return "combination_grid"
    if hits(["murder", "manslaughter", "felony murder", "depraved"]) >= 2:
        return "homicide_ladder"
    if hits(["larceny", "embezzlement", "false pretenses", "larceny by trick", "robbery", "burglary", "arson", "forgery"]) >= 2:
        return "property_label"
    if hits(["attempt", "conspiracy", "solicitation", "merge", "withdraw"]) >= 2:
        return "inchoate_scope"
    if hits(["suppress", "admissible", "inadmissible", "reversed", "affirmed", "harmless", "grant", "deny"]) >= 2:
        return "remedy_result"
    if hits(["fifth amendment", "sixth amendment", "due process", "right to counsel", "miranda", "confrontation"]) >= 2:
        return "right_source"
    if hits(["warrant", "probable cause", "reasonable suspicion", "search incident", "exigent", "plain view", "particularity"]) >= 2:
        return "search_threshold"
    if "yes, because" in text and "no, because" in text:
        return "yes_no_because"
    if "guilty" in text and "not guilty" in text:
        return "guilty_not_guilty"
    return "other"


def pick_longest_words(rec):
    counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
    return max(LETTERS, key=lambda l: counts[l])


def pick_longest_chars(rec):
    return max(LETTERS, key=lambda l: len(rec["choices"][l]))


def pick_shortest_words(rec):
    counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
    return min(LETTERS, key=lambda l: counts[l])


def pick_position(letter: str):
    return lambda rec: letter


def pick_top2_len_pref(letter: str):
    def fn(rec):
        counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
        top = sorted(LETTERS, key=lambda l: (-counts[l], LETTERS.index(l)))[:2]
        return letter if letter in top else top[0]
    return fn


def pick_top3_len_pref(letter: str):
    def fn(rec):
        counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
        top = sorted(LETTERS, key=lambda l: (-counts[l], LETTERS.index(l)))[:3]
        return letter if letter in top else top[0]
    return fn


def pick_unique_longest_else(letter: str):
    def fn(rec):
        counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
        max_len = max(counts.values())
        cands = [l for l in LETTERS if counts[l] == max_len]
        return cands[0] if len(cands) == 1 else letter
    return fn


def pick_longest_abc_else_c(rec):
    counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
    max_len = max(counts.values())
    cands = [l for l in LETTERS if counts[l] == max_len]
    if len(cands) == 1 and cands[0] in "ABC":
        return cands[0]
    return "C"


def pick_longest_set_empirical(rec):
    counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
    max_len = max(counts.values())
    cands = tuple(l for l in LETTERS if counts[l] == max_len)
    mapping = {
        ("A",): "A",
        ("B",): "B",
        ("C",): "C",
        ("D",): "C",
        ("A", "B"): "B",
        ("A", "C"): "C",
        ("A", "D"): "C",
        ("B", "C"): "A",
        ("B", "D"): "C",
        ("C", "D"): "C",
        ("A", "B", "C"): "A",
        ("B", "C", "D"): "C",
        ("A", "B", "C", "D"): "A",
    }
    return mapping.get(cands, "C")


def pick_most_legal_terms_else_longest(rec):
    scores = {}
    for l in LETTERS:
        text = rec["choices"][l]
        scores[l] = len(REMEDY_RE.findall(text)) + len(RIGHT_RE.findall(text)) + len(CRIME_RE.findall(text))
    best = max(scores.values())
    cands = [l for l in LETTERS if scores[l] == best]
    if len(cands) == 1 and best > 0:
        return cands[0]
    return pick_longest_words(rec)


def evaluate(records, fn):
    hits = 0
    for rec in records:
        pred = fn(rec)
        if pred == rec["correct"]:
            hits += 1
    return hits / len(records) if records else 0.0


def bootstrap_ci(records, fn, n=1000, seed=13):
    rng = random.Random(seed)
    scores = []
    for _ in range(n):
        sample = [records[rng.randrange(len(records))] for _ in records]
        scores.append(evaluate(sample, fn))
    scores.sort()
    return scores[int(0.025 * n)], scores[int(0.975 * n)]


def fold_indices(n, k=5):
    return {i: i % k for i in range(n)}


def cv_letter_rule(records, key_fn):
    folds = fold_indices(len(records), 5)
    hits = 0
    for fold in range(5):
        train = [r for i, r in enumerate(records) if folds[i] != fold]
        test = [r for i, r in enumerate(records) if folds[i] == fold]
        buckets = defaultdict(Counter)
        global_counts = Counter(r["correct"] for r in train)
        for r in train:
            buckets[key_fn(r)][r["correct"]] += 1
        default = global_counts.most_common(1)[0][0]
        for r in test:
            counts = buckets.get(key_fn(r))
            pred = counts.most_common(1)[0][0] if counts else default
            if pred == r["correct"]:
                hits += 1
    return hits / len(records)


def cv_best_rule_by_menu(records, candidate_rules):
    folds = fold_indices(len(records), 5)
    hits = 0
    chosen_by_fold = []
    for fold in range(5):
        train = [r for i, r in enumerate(records) if folds[i] != fold]
        test = [r for i, r in enumerate(records) if folds[i] == fold]
        menus = sorted(set(menu_type(r["choices"]) for r in train))
        selected = {}
        for menu in menus:
            group = [r for r in train if menu_type(r["choices"]) == menu]
            if len(group) < 8:
                selected[menu] = "longest_A/B/C_else_C"
                continue
            scored = []
            for rule_name, rule_fn in candidate_rules.items():
                scored.append((evaluate(group, rule_fn), rule_name))
            scored.sort(reverse=True)
            selected[menu] = scored[0][1]
        chosen_by_fold.append(selected)
        for rec in test:
            menu = menu_type(rec["choices"])
            rule_name = selected.get(menu, "longest_A/B/C_else_C")
            pred = candidate_rules[rule_name](rec)
            if pred == rec["correct"]:
                hits += 1
    return hits / len(records), chosen_by_fold


def sklearn_cv(records):
    feature_names = [
        "word_count", "char_count", "rel_word_len", "rel_char_len",
        "is_longest_words", "is_longest_chars", "word_length_rank", "char_length_rank",
        "qualifiers", "absolutes", "negations", "remedy_words", "right_words", "crime_words",
        "has_because", "starts_yes", "starts_no", "has_guilty", "has_not_guilty", "has_only",
        "position",
    ]
    try:
        from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
        from sklearn.linear_model import LogisticRegression
        from sklearn.pipeline import make_pipeline
        from sklearn.preprocessing import StandardScaler
    except Exception as exc:
        return {"available": False, "error": str(exc)}

    models = {
        "logistic": make_pipeline(StandardScaler(), LogisticRegression(max_iter=2000, C=0.7)),
        "random_forest": RandomForestClassifier(n_estimators=300, max_depth=5, random_state=13, class_weight="balanced_subsample"),
        "gradient_boosting": GradientBoostingClassifier(random_state=13, max_depth=2, n_estimators=120, learning_rate=0.06),
    }
    folds = fold_indices(len(records), 5)
    results = {}
    for model_name, model in models.items():
        hits = 0
        for fold in range(5):
            train_q = [i for i in range(len(records)) if folds[i] != fold]
            test_q = [i for i in range(len(records)) if folds[i] == fold]
            X, y = [], []
            for qi in train_q:
                rec = records[qi]
                for letter in LETTERS:
                    X.append(feature_vector(rec["choices"][letter], letter, rec["choices"], feature_names))
                    y.append(1 if rec["correct"] == letter else 0)
            model.fit(X, y)
            for qi in test_q:
                rec = records[qi]
                candidates = [feature_vector(rec["choices"][letter], letter, rec["choices"], feature_names) for letter in LETTERS]
                probs = model.predict_proba(candidates)[:, 1]
                pred = LETTERS[max(range(4), key=lambda i: probs[i])]
                if pred == rec["correct"]:
                    hits += 1
        results[model_name] = hits / len(records)
    return {"available": True, "feature_names": feature_names, "results": results}


def topn_contains(records, n):
    hits = 0
    for rec in records:
        counts = {l: word_count(rec["choices"][l]) for l in LETTERS}
        top = sorted(LETTERS, key=lambda l: (-counts[l], LETTERS.index(l)))[:n]
        if rec["correct"] in top:
            hits += 1
    return hits / len(records)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    records = load_records()
    graded = [r for r in records if r["percent_correct"] is not None]
    critical = [r for r in records if r["percent_correct"] is not None and r["percent_correct"] < 50]
    easy = [r for r in records if r["percent_correct"] is not None and r["percent_correct"] >= 65]

    rules = {
        "always_A": pick_position("A"),
        "always_B": pick_position("B"),
        "always_C": pick_position("C"),
        "always_D": pick_position("D"),
        "longest_words": pick_longest_words,
        "longest_chars": pick_longest_chars,
        "shortest_words": pick_shortest_words,
        "unique_longest_else_C": pick_unique_longest_else("C"),
        "longest_A/B/C_else_C": pick_longest_abc_else_c,
        "longest_set_empirical_map": pick_longest_set_empirical,
        "C_if_top2_length_else_longest": pick_top2_len_pref("C"),
        "C_if_top3_length_else_longest": pick_top3_len_pref("C"),
        "B_if_top2_length_else_longest": pick_top2_len_pref("B"),
        "A_if_top2_length_else_longest": pick_top2_len_pref("A"),
        "D_if_top2_length_else_longest": pick_top2_len_pref("D"),
        "most_legal_terms_else_longest": pick_most_legal_terms_else_longest,
    }

    rule_rows = []
    for name, fn in rules.items():
        lo, hi = bootstrap_ci(records, fn, n=500)
        rule_rows.append(
            {
                "rule": name,
                "all_accuracy": evaluate(records, fn),
                "all_ci_low": lo,
                "all_ci_high": hi,
                "graded_accuracy": evaluate(graded, fn),
                "critical_accuracy": evaluate(critical, fn),
                "easy_accuracy": evaluate(easy, fn),
            }
        )
    rule_rows.sort(key=lambda x: x["all_accuracy"], reverse=True)

    menu_rows = []
    for menu, group in defaultdict(list, {}).items():
        pass
    grouped = defaultdict(list)
    for rec in records:
        grouped[menu_type(rec["choices"])].append(rec)
    for menu, group in grouped.items():
        menu_rows.append(
            {
                "menu_type": menu,
                "n": len(group),
                "longest_words_accuracy": evaluate(group, pick_longest_words),
                "C_if_top2_length_else_longest_accuracy": evaluate(group, pick_top2_len_pref("C")),
                "always_C_accuracy": evaluate(group, pick_position("C")),
                "top2_length_contains_correct": topn_contains(group, 2),
                "top3_length_contains_correct": topn_contains(group, 3),
            }
        )
    menu_rows.sort(key=lambda x: x["n"], reverse=True)

    cv_results = {
        "best_letter_by_menu_type": cv_letter_rule(records, lambda r: menu_type(r["choices"])),
        "best_letter_by_menu_and_longest_letter": cv_letter_rule(records, lambda r: (menu_type(r["choices"]), pick_longest_words(r))),
        "best_letter_by_longest_letter": cv_letter_rule(records, lambda r: pick_longest_words(r)),
        "sklearn": sklearn_cv(records),
    }
    menu_rule_candidates = {
        name: fn for name, fn in rules.items()
        if name in {
            "always_A", "always_B", "always_C", "always_D", "longest_words",
            "longest_chars", "unique_longest_else_C", "longest_A/B/C_else_C",
            "C_if_top2_length_else_longest", "B_if_top2_length_else_longest",
            "A_if_top2_length_else_longest", "D_if_top2_length_else_longest",
        }
    }
    menu_rule_acc, menu_rule_choices = cv_best_rule_by_menu(records, menu_rule_candidates)
    cv_results["best_candidate_rule_by_menu_type"] = menu_rule_acc
    cv_results["best_candidate_rule_by_menu_type_choices"] = menu_rule_choices

    summary = {
        "source": str(SRC),
        "n_records": len(records),
        "n_graded": len(graded),
        "n_critical_below_50": len(critical),
        "correct_letter_counts": dict(Counter(r["correct"] for r in records)),
        "popular_wrong_counts": dict(Counter(r["popular_wrong"] for r in records if r["popular_wrong"] in LETTERS)),
        "top2_length_contains_correct": topn_contains(records, 2),
        "top3_length_contains_correct": topn_contains(records, 3),
        "rules": rule_rows,
        "menu_breakdown": menu_rows,
        "cv_results": cv_results,
    }

    (OUT / "mechanical_choices_only_results.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")

    with (OUT / "mechanical_choices_only_rule_scoreboard.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rule_rows[0].keys()))
        writer.writeheader()
        writer.writerows(rule_rows)

    with (OUT / "mechanical_choices_only_menu_breakdown.csv").open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(menu_rows[0].keys()))
        writer.writeheader()
        writer.writerows(menu_rows)

    print(json.dumps({
        "n_records": summary["n_records"],
        "top_rules": rule_rows[:8],
        "top2_length_contains_correct": summary["top2_length_contains_correct"],
        "top3_length_contains_correct": summary["top3_length_contains_correct"],
        "cv_results": cv_results,
        "top_menus": menu_rows[:8],
    }, indent=2))


if __name__ == "__main__":
    main()
